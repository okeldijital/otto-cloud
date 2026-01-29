#!/usr/bin/env python3
import openpyxl
import json

# Load the spreadsheet
wb = openpyxl.load_workbook('/home/nkululeko/Desktop/Music Label Management Spreadsheet.xlsx')

schema = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Get headers (first row)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    # Get sample data (first 3 rows after header)
    sample_data = []
    for row_idx in range(2, min(5, ws.max_row + 1)):
        row_data = {}
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[header] = str(cell.value) if cell.value is not None else None
        sample_data.append(row_data)
    
    schema[sheet_name] = {
        'columns': headers,
        'row_count': ws.max_row - 1,  # Excluding header
        'sample_data': sample_data
    }

# Print the schema
print(json.dumps(schema, indent=2))
